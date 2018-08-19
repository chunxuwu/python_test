# -*- coding: utf-8 -*-
"""
Created on Sun Aug 19 13:07:23 2018
12.10 调整行和列

1、设置高度和宽度
2、合并和拆分单元格

3、冻结窗口，冻结顶部几行或最左边几列，对阅读很有帮助


@author: NEVERGUVEIP
"""

import openpyxl
wb = openpyxl.Workbook()

sheet = wb.get_active_sheet()

sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'
##设置宽和高
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20


##和并拆分单元格
sheet.merge_cells('A1:D3')
sheet['A1'] = 'twelve cells merged together'
sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two merged cells.'

##差分单元格
sheet.unmerge_cells('A1:D3')



wb.save('dimensions.xlsx')
