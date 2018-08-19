# -*- coding: utf-8 -*-
"""
Created on Sun Aug 19 10:05:23 2018

12.6 更新电子表格
从单元表中找到特定类型的产品，并更新他们的价格

12.7 设置单元格的字体风格

12.9 公式
关于
https://blog.csdn.net/weixin_41569319/article/details/80807006
关于没有style模块的帖子

对于表格，如果希望看到计算结果，不想看其中的公式，必须将load_workbook()对象的data_only设置
为True
@author: NEVERGUVEIP
"""
#！python3

import openpyxl
from openpyxl.styles import Font

wb = openpyxl.load_workbook(r'.\automate_online-materials\produceSales.xlsx',data_only=True)
sheet = wb.get_sheet_by_name('Sheet')

#需要更新的价格
PRICE_UPDATES = {'Garlic':3.07,
                 'Celery':1.19,
                 'Lemon':1.27}


italic24Font = Font(size = 24,italic = True)

c = (sheet.cell(row = 2,column = 4).value)
d = sheet.cell(row = 2,column = 3).value
print(c)##
print(type(d))
#更新所有行，更新不正确的价格
for rowNum in range(2,sheet.max_row):
    produceName = sheet.cell(row=rowNum,column=1).value
    if produceName in PRICE_UPDATES:
        #设置单元格的风格
        sheet.cell(row = rowNum,column = 1).font = italic24Font
        sheet.cell(row = rowNum,column = 2).font = italic24Font
        sheet.cell(row = rowNum,column = 2).value = PRICE_UPDATES[produceName]
        #sheet.cell(row = rowNum,column = 5).value = '=SUM(B2:D2)'
        a = float(sheet.cell(row = rowNum,column = 2).value)
        b = float(sheet.cell(row = rowNum,column = 3).value)
        c = float(sheet.cell(row = rowNum,column = 4).value)
        #为啥c不能转换成float类型呢？
        #没有设置成data_only为True
        
        sheet.cell(row = rowNum,column = 5).value = str(a+b+c)

wb.save('updateProduceSales.xlsx')