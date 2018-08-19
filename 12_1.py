# -*- coding: utf-8 -*-
"""
Created on Sat Aug 18 14:47:02 2018


根据属性取得表格中的值
根据行列数求取表格中的值

@author: NEVERGUVEIP
"""
import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
print(type(wb))

print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('Sheet1')#大写的Sheet3
print(sheet.title)

print(sheet['A1'].value)
print(sheet['B1'].value)##取得表格sheet1中B1的值

c = sheet['B1']
print('Row'+str(c.row)+':Column'+ c.column + ' is\n' + c.value)

#coordinate 给出的属性是‘B1’
print('Cell '+c.coordinate+' is '+ c.value)

print(sheet.cell(row=1,column=2))

for i in range(1,8,2):
    print(i,sheet.cell(row=i,column=1).value)

