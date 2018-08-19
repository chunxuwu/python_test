# -*- coding: utf-8 -*-
"""
Created on Sat Aug 18 15:56:16 2018
总结：
1、导入openpyxl模块
2、调用openpyxl.load_workbook()函数
3、取得Workbook对象
4、调用get_active_sheet()或get_sheet_by_name()工作簿方法
5、取得Worksheet对象
6、使用索引或工作表的cell()方法，带上row和column关键字参数
7、取得Cell对象
8、读取Cell对象的value属性



@author: NEVERGUVEIP
"""

import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string

print(get_column_letter(1))#获取第一列的字母

print(get_column_letter(900))


wb = openpyxl.load_workbook('example.xlsx')

sheet = wb.get_sheet_by_name('Sheet1')#读取sheet1表格

print(get_column_letter(sheet.max_column))#得到最大列的字母

a = column_index_from_string('AA')#根据列字母转换成数字

print(a)

#在一个元组中列出cell对象
print(tuple(sheet['A1':'C3']))

#遍历表格中的内容，使用俩个for循环
for cellObject in sheet['A1':'C3']:
    for cellobj in cellObject:
        print(cellobj.coordinate,cellobj.value)
    print('-------end-------')