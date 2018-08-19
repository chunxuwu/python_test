# -*- coding: utf-8 -*-
"""
Created on Sun Aug 19 09:08:00 2018

12.5 写入Excle文档
@author: NEVERGUVEIP
"""

import openpyxl
wb = openpyxl.Workbook()#创建一个名为sheet的表格
#print(wb.get_sheet_names())
#默认新工作表是添加在工作簿最后，也可以通过索引指定位置
wb.create_sheet()#添加一个新的工作表sheet1
wb.create_sheet(index=2,title='middle sheet')
wb.create_sheet(index=3,title='Sheet2')

sheet = wb.get_active_sheet()#激活第一个工作表吗？
print(sheet.title)
sheet.title = 'Spam Bacon Eggs Sheet'

#向指定的工作表写入数据
sheet1=wb.get_sheet_by_name('Sheet2')
sheet1['A1'] = 'Hello world!'



print(wb.get_sheet_names())


#通过工作表的名字删除工作表
wb.remove_sheet(wb.get_sheet_by_name('Sheet1'))
print(wb.get_sheet_names())
wb.save('example_copy.xlsx')

