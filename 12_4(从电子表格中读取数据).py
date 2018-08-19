# -*- coding: utf-8 -*-
"""
Created on Sat Aug 18 21:40:01 2018

@author: NEVERGUVEIP
"""

#！ python3
import openpyxl,pprint
print('Opening workbook...')

#wb = openpyxl.load_workbook(r'C:\Users\NEVERGUVEIP\Documents\GitHub\python_test\automate_online-materials\censuspopdata.xlsx')
wb = openpyxl.load_workbook(r'.\automate_online-materials\censuspopdata.xlsx')

sheet = wb.get_sheet_by_name('Population by Census Tract')

countyData = {}

#TODO:fill in countyData with each county's population and ttacts
print('rasding row...')

for row in range(2,sheet.max_row+1):
    #each row in the speadsheet has data for one census tract
    state = sheet['B'+str(row)].value
    county = sheet['C'+str(row)].value
    pop = sheet['D'+str(row)].value
    #make sure the key for this state exists.
    
    #往字典中写入数据
    countyData.setdefault(state,{})
    
    countyData[state].setdefault(county,{'tracts':0,'pop':0})
    
    #each row represents one census tract,so increment tract
    countyData[state][county]['tracts'] += 1
    
    #increse the county pop by the pop in the census tract.
    countyData[state][county]['pop']+=int(pop)

#将结果写入到py文件中的好处是
#可以像导入其他模块一样导入py文件
print('writing results...')
resultFile = open('census2010.py','w')
resultFile.write('allData = '+pprint.pformat(countyData))
resultFile.close()
print('Done.')

#读取census2010中的数据
#import os
#os.chdir('')
import census2010
print(census2010.allData['AK']['Anchorage'])
anchoragePop = census2010.allData['AK']['Anchorage']['pop']
print('the 2010 population of Anchorage was '+ str(anchoragePop))


    